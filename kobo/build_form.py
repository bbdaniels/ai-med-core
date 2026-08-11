#!/usr/bin/env python3
"""Build XLSForm (.xlsx) files from JSON template definitions.

Usage:
    python kobo/build_form.py                                # Build all templates
    python kobo/build_form.py kobo/templates/diagnosis.json  # Build one form
    python kobo/build_form.py --list                         # List available templates
"""

import json
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook

KOBO_DIR = Path(__file__).parent
TEMPLATES_DIR = KOBO_DIR / "templates"
OUTPUT_DIR = KOBO_DIR / "forms"


def collect_languages(defn: dict) -> list[str]:
    """Discover all languages used in the form definition.

    Returns a sorted list of language keys (e.g. ["English (en)", "Français (fr)"]).
    If no multi-language labels are found, returns an empty list.
    """
    langs: set[str] = set()
    for field in defn.get("fields", []):
        label = field.get("label")
        if isinstance(label, dict):
            langs.update(label.keys())
    # Also check choices
    for options in defn.get("choices", {}).values():
        for opt in options:
            label = opt.get("label")
            if isinstance(label, dict):
                langs.update(label.keys())
    if not langs:
        return []
    # Sort with default language first
    default_lang = defn.get("settings", {}).get("default_language", "English (en)")
    rest = sorted(langs - {default_lang})
    if default_lang in langs:
        return [default_lang] + rest
    return rest


def build_form(definition_path: Path) -> Path:
    """Build an XLSForm from a JSON form definition."""
    with open(definition_path) as f:
        defn = json.load(f)

    languages = collect_languages(defn)
    multi_lang = len(languages) > 0

    wb = Workbook()

    # === SURVEY SHEET ===
    survey = wb.active
    survey.title = "survey"

    if multi_lang:
        label_cols = [f"label::{lang}" for lang in languages]
        survey.append(["type", "name"] + label_cols + ["required", "appearance", "default", "relevant", "parameters"])
    else:
        survey.append(["type", "name", "label", "required", "appearance", "default", "relevant", "parameters"])

    # Auto timestamps
    empty_labels = [""] * len(languages) if multi_lang else [""]
    survey.append(["start", "start"] + empty_labels + ["", "", "", "", ""])
    survey.append(["end", "end"] + empty_labels + ["", "", "", "", ""])

    # Fields
    for field in defn["fields"]:
        ftype = field["type"]
        name = field["name"]
        label_raw = field.get("label", "")
        required = "yes" if field.get("required", False) else ""
        appearance = field.get("appearance", "")
        default = field.get("default", "")
        relevant = field.get("relevant", "")
        parameters = field.get("parameters", "")

        # Build label columns
        if multi_lang:
            if isinstance(label_raw, dict):
                labels = [label_raw.get(lang, "") for lang in languages]
            else:
                # Single string label — put in all language columns
                labels = [label_raw] * len(languages)
        else:
            labels = [label_raw if isinstance(label_raw, str) else str(label_raw)]

        if ftype == "note":
            survey.append(["note", name] + labels + ["", "", "", relevant, ""])
        elif ftype == "text":
            survey.append(["text", name] + labels + [required, appearance, default, relevant, ""])
        elif ftype == "integer":
            survey.append(["integer", name] + labels + [required, appearance, default, relevant, ""])
        elif ftype == "range":
            survey.append(["range", name] + labels + [required, appearance, default, relevant, parameters])
        elif ftype == "select_one":
            list_name = field["list"]
            survey.append([f"select_one {list_name}", name] + labels + [required, appearance, default, relevant, ""])
        elif ftype == "select_multiple":
            list_name = field["list"]
            survey.append([f"select_multiple {list_name}", name] + labels + [required, appearance, default, relevant, ""])
        elif ftype == "hidden":
            hidden_label = labels if multi_lang else [label_raw or name]
            survey.append(["hidden", name] + hidden_label + ["", "", default, "", ""])
        elif ftype == "begin_group":
            survey.append(["begin_group", name] + labels + ["", appearance, "", relevant, ""])
        elif ftype == "end_group":
            survey.append(["end_group", ""] + ([""] * len(languages) if multi_lang else [""]) + ["", "", "", "", ""])
        else:
            survey.append([ftype, name] + labels + [required, appearance, default, relevant, parameters])

    # === CHOICES SHEET ===
    choices_data = defn.get("choices", {})
    if choices_data:
        choices = wb.create_sheet("choices")
        if multi_lang:
            choice_label_cols = [f"label::{lang}" for lang in languages]
            choices.append(["list_name", "name"] + choice_label_cols)
            for list_name, options in choices_data.items():
                for opt in options:
                    opt_label = opt.get("label", "")
                    if isinstance(opt_label, dict):
                        opt_labels = [opt_label.get(lang, "") for lang in languages]
                    else:
                        opt_labels = [opt_label] * len(languages)
                    choices.append([list_name, opt["name"]] + opt_labels)
        else:
            choices.append(["list_name", "name", "label"])
            for list_name, options in choices_data.items():
                for opt in options:
                    choices.append([list_name, opt["name"], opt["label"]])

    # === SETTINGS SHEET ===
    settings_data = defn.get("settings", {})
    settings = wb.create_sheet("settings")
    form_title_raw = settings_data.get("form_title", defn.get("name", "Untitled"))
    form_id = settings_data.get("form_id", definition_path.stem)
    version = settings_data.get("version", "1")
    default_language = settings_data.get("default_language", "English (en)")

    if isinstance(form_title_raw, dict) and multi_lang:
        # Multi-language form title: form_title::Language (code) columns
        title_cols = [f"form_title::{lang}" for lang in languages]
        settings.append(title_cols + ["form_id", "version", "default_language"])
        title_values = [form_title_raw.get(lang, "") for lang in languages]
        settings.append(title_values + [form_id, version, default_language])
    else:
        title_str = form_title_raw if isinstance(form_title_raw, str) else str(form_title_raw)
        settings.append(["form_title", "form_id", "version", "default_language"])
        settings.append([title_str, form_id, version, default_language])

    # Save — output next to source file if outside templates dir, else to forms/
    if definition_path.resolve().parent == TEMPLATES_DIR.resolve():
        out_dir = OUTPUT_DIR
    else:
        out_dir = definition_path.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    output_path = out_dir / f"{definition_path.stem}.xlsx"
    wb.save(output_path)
    return output_path


def main():
    if len(sys.argv) > 1 and sys.argv[1] == "--list":
        forms = sorted(TEMPLATES_DIR.glob("*.json"))
        if not forms:
            print("No form definitions found in kobo/templates/")
            return
        print(f"Found {len(forms)} form definition(s):")
        for f in forms:
            with open(f) as fh:
                d = json.load(fh)
            print(f"  {f.name:30s}  {d.get('name', '(unnamed)')}")
        return

    if len(sys.argv) > 1:
        # Build specific form
        path = Path(sys.argv[1])
        if not path.exists():
            print(f"Error: {path} not found")
            sys.exit(1)
        out = build_form(path)
        print(f"Built: {out}")
    else:
        # Build all forms
        forms = sorted(TEMPLATES_DIR.glob("*.json"))
        if not forms:
            print("No form definitions found in kobo/templates/")
            return
        for form_path in forms:
            out = build_form(form_path)
            print(f"Built: {out}")


if __name__ == "__main__":
    main()
